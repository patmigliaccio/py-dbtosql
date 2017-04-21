import openpyxl

class Excel(object):

	def __init__(self, filePath):
		self.filePath = filePath

	def open(self):
		self.workbook = openpyxl.load_workbook(self.filePath)

	def retrieve(self, sheet_index = 0):
		self.data = []
		self.sheet = self.workbook.worksheets[sheet_index]
		return self.retrieve_data()

	def retrieve_data(self):
		rows = self.sheet.max_row
		cols = self.sheet.max_column

		data = []
		headers = []

		for row in range(1, rows + 1):
			rowData = {}

			for col in range(1, cols + 1):
				cell = self.sheet.cell(row = row, column = col).value
				
				# adds the first row as header values
				if row == 1:
					headers.append(cell)
				else:	
					if cell != None:
						# creates a hashtable of data for the rows and columns
						rowData[headers[col - 1]] = cell

			if row != 1:
				data.append(rowData)

		return data




		